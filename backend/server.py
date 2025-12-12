from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
from jose import JWTError, jwt
import requests
import base64
from io import BytesIO
from PIL import Image
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()
SECRET_KEY = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_DAYS = 7

# Holdprint API Keys
HOLDPRINT_API_KEY_POA = os.environ.get('HOLDPRINT_API_KEY_POA')
HOLDPRINT_API_KEY_SP = os.environ.get('HOLDPRINT_API_KEY_SP')
HOLDPRINT_API_URL = "https://api.holdworks.ai/api-key/jobs/data"

# ============ CATÁLOGO DE PRODUTOS HOLDPRINT ============
# Mapeamento de produtos para famílias - usado para associação automática

PRODUCT_FAMILY_MAPPING = {
    # Adesivos
    "Adesivos": [
        "adesivo", "vinil", "fachada adesivada", "fachada com vinil"
    ],
    # Lonas e Banners
    "Lonas e Banners": [
        "lona", "banner", "faixa", "empena", "faixa de gradil"
    ],
    # Chapas e Placas
    "Chapas e Placas": [
        "chapa", "placa", "acm", "acrílico", "mdf", "ps", "pvc", "polionda", 
        "policarbonato", "petg", "compensado", "xps"
    ],
    # Estruturas Metálicas
    "Estruturas Metálicas": [
        "estrutura metálica", "estrutura metalica", "backdrop", "cavalete"
    ],
    # Tecidos
    "Tecidos": [
        "tecido", "bandeira", "wind banner"
    ],
    # Letras Caixa
    "Letras Caixa": [
        "letra caixa", "letra-caixa", "letras caixa"
    ],
    # Totens
    "Totens": [
        "totem"
    ],
    # Envelopamento
    "Envelopamento": [
        "envelopamento"
    ],
    # Painéis Luminosos
    "Painéis Luminosos": [
        "painel backlight", "painel luminoso", "backlight"
    ],
    # Serviços
    "Serviços": [
        "serviço", "serviços", "instalação", "entrega", "montagem", 
        "pintura", "serralheria", "solda", "corte", "aplicação"
    ],
    # Materiais Promocionais
    "Materiais Promocionais": [
        "cartaz", "flyer", "folder", "panfleto", "imã", "marca-página"
    ],
    # Produtos Terceirizados
    "Produtos Terceirizados": [
        "terceirizado", "produto genérico"
    ],
    # Sublimação
    "Sublimação": [
        "sublimação", "sublimática", "sublimatico"
    ],
    # Impressão
    "Impressão": [
        "impressão uv", "impressão latex", "impressão solvente"
    ],
    # Display/PS
    "Display/PS": [
        "display", "móbile", "mobile", "orelha de monitor"
    ],
    # Fundação
    "Fundação/Estrutura": [
        "fundação", "sapata", "estrutura em madeira"
    ]
}

def classify_product_to_family(product_name: str) -> tuple:
    """
    Classifica um produto em uma família baseado no nome.
    Retorna (family_name, confidence_score)
    """
    if not product_name:
        return (None, 0)
    
    product_lower = product_name.lower()
    
    # Mapeamento com prioridade (mais específico primeiro)
    priority_mapping = [
        # Letras Caixa - verificar antes de outros
        ("Letras Caixa", ["letra caixa", "letra-caixa", "letras caixa"]),
        # Totens
        ("Totens", ["totem"]),
        # Envelopamento
        ("Envelopamento", ["envelopamento", "envelopar"]),
        # Painéis Luminosos
        ("Painéis Luminosos", ["painel backlight", "painel luminoso", "backlight", "lightbox"]),
        # Tecidos
        ("Tecidos", ["tecido", "bandeira", "wind banner"]),
        # Estruturas Metálicas
        ("Estruturas Metálicas", ["estrutura metálica", "estrutura metalica", "backdrop", "cavalete"]),
        # Lonas e Banners
        ("Lonas e Banners", ["lona", "banner", "faixa", "empena"]),
        # Adesivos - depois de lonas para não pegar "lona com adesivo"
        ("Adesivos", ["adesivo", "vinil", "fachada adesivada", "fachada com vinil"]),
        # Chapas e Placas
        ("Chapas e Placas", ["chapa", "placa", "acm", "acrílico", "acrilico", "mdf", " ps ", "pvc", "polionda", 
                           "policarbonato", "petg", "compensado", "xps"]),
        # Serviços
        ("Serviços", ["serviço", "serviços", "instalação", "instalacao", "entrega", "montagem", 
                     "pintura", "serralheria", "solda", "corte", "aplicação", "aplicacao"]),
        # Materiais Promocionais
        ("Materiais Promocionais", ["cartaz", "flyer", "folder", "panfleto", "imã", "marca-página"]),
        # Sublimação
        ("Sublimação", ["sublimação", "sublimática", "sublimatico", "sublimacao"]),
        # Impressão
        ("Impressão", ["impressão uv", "impressão latex", "impressão solvente", "impresso"]),
        # Display/PS
        ("Display/PS", ["display", "móbile", "mobile", "orelha de monitor"]),
        # Produtos Terceirizados
        ("Produtos Terceirizados", ["terceirizado", "produto genérico"]),
        # Fundação
        ("Fundação/Estrutura", ["fundação", "sapata", "estrutura em madeira"]),
    ]
    
    best_match = None
    best_score = 0
    
    for family_name, keywords in priority_mapping:
        for keyword in keywords:
            if keyword.lower() in product_lower:
                # Score baseado no tamanho do match e posição
                keyword_len = len(keyword)
                product_len = len(product_name)
                
                # Score base: proporção do keyword no nome
                base_score = (keyword_len / product_len) * 100
                
                # Bonus se keyword está no início
                if product_lower.startswith(keyword.lower()):
                    base_score += 30
                
                # Bonus por match exato de palavra
                if keyword.lower() == product_lower:
                    base_score = 100
                
                score = min(base_score, 100)
                
                if score > best_score:
                    best_score = score
                    best_match = family_name
    
    if best_match:
        return (best_match, round(best_score, 1))
    
    return ("Outros", 10)  # Família genérica com baixa confiança

def extract_product_measures(description: str) -> dict:
    """
    Extrai medidas (largura, altura, cópias) da descrição HTML do produto.
    Retorna dict com width_m, height_m, copies e area_m2
    """
    import re
    
    result = {
        "width_m": None,
        "height_m": None,
        "copies": 1,
        "area_m2": None
    }
    
    if not description:
        return result
    
    # Extrair Largura - vários formatos possíveis
    width_patterns = [
        r'Largura:\s*<span[^>]*>([0-9.,]+)\s*m',
        r'Largura:\s*([0-9.,]+)\s*m',
        r'largura[:\s]+([0-9.,]+)\s*m',
    ]
    for pattern in width_patterns:
        match = re.search(pattern, description, re.IGNORECASE)
        if match:
            result["width_m"] = float(match.group(1).replace(',', '.'))
            break
    
    # Extrair Altura
    height_patterns = [
        r'Altura:\s*<span[^>]*>([0-9.,]+)\s*m',
        r'Altura:\s*([0-9.,]+)\s*m',
        r'altura[:\s]+([0-9.,]+)\s*m',
    ]
    for pattern in height_patterns:
        match = re.search(pattern, description, re.IGNORECASE)
        if match:
            result["height_m"] = float(match.group(1).replace(',', '.'))
            break
    
    # Extrair Cópias
    copies_patterns = [
        r'Cópias:\s*<span[^>]*>([0-9]+)',
        r'Cópias:\s*([0-9]+)',
        r'copias[:\s]+([0-9]+)',
    ]
    for pattern in copies_patterns:
        match = re.search(pattern, description, re.IGNORECASE)
        if match:
            result["copies"] = int(match.group(1))
            break
    
    # Calcular área se tiver largura e altura
    if result["width_m"] and result["height_m"]:
        result["area_m2"] = round(result["width_m"] * result["height_m"] * result["copies"], 2)
    
    return result

def calculate_job_products_area(holdprint_data: dict) -> tuple:
    """
    Calcula a área de todos os produtos de um job.
    Retorna (products_with_area, total_area_m2, total_products, total_quantity)
    """
    products = holdprint_data.get("products", [])
    products_with_area = []
    total_area_m2 = 0
    total_quantity = 0
    
    for product in products:
        product_name = product.get("name", "")
        quantity = product.get("quantity", 1)
        description = product.get("description", "")
        
        # Extrair medidas
        measures = extract_product_measures(description)
        
        # Classificar família
        family_name, confidence = classify_product_to_family(product_name)
        
        # Calcular área do item (considerando quantidade)
        item_area = None
        if measures["width_m"] and measures["height_m"]:
            # Área unitária × quantidade
            unit_area = measures["width_m"] * measures["height_m"]
            item_area = round(unit_area * quantity * measures["copies"], 2)
            total_area_m2 += item_area
        
        total_quantity += quantity
        
        product_data = {
            "name": product_name,
            "family_name": family_name,
            "confidence": confidence,
            "quantity": quantity,
            "width_m": measures["width_m"],
            "height_m": measures["height_m"],
            "copies": measures["copies"],
            "unit_area_m2": round(measures["width_m"] * measures["height_m"], 2) if measures["width_m"] and measures["height_m"] else None,
            "total_area_m2": item_area,
            "unit_price": product.get("unitPrice", 0),
            "total_value": product.get("totalValue", 0)
        }
        products_with_area.append(product_data)
    
    return (products_with_area, round(total_area_m2, 2), len(products), total_quantity)

app = FastAPI()
api_router = APIRouter(prefix="/api")

# ============ MODELS ============

class UserRole:
    ADMIN = "admin"
    MANAGER = "manager"
    INSTALLER = "installer"

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    name: str
    role: str = UserRole.INSTALLER
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    is_active: bool = True

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = UserRole.INSTALLER

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class Installer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    full_name: str
    phone: Optional[str] = None
    branch: str  # POA or SP
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Job(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    holdprint_job_id: str
    title: str
    client_name: str
    client_address: Optional[str] = None
    status: str = "aguardando"  # aguardando, instalando, pausado, finalizado, atrasado
    area_m2: Optional[float] = None  # Área total calculada do job
    branch: str  # POA or SP
    assigned_installers: List[str] = []  # List of installer IDs
    scheduled_date: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    items: List[dict] = []  # Job items from Holdprint
    holdprint_data: dict = {}  # Raw data from Holdprint
    # Campos calculados para análise de produtividade
    products_with_area: List[dict] = []  # Produtos com área calculada
    total_products: int = 0
    total_quantity: int = 0
    # Atribuição de itens a instaladores
    item_assignments: List[dict] = []  # [{item_index, installer_id, installer_name, assigned_at}]

class JobCreate(BaseModel):
    holdprint_job_id: str
    branch: str

class JobAssign(BaseModel):
    installer_ids: List[str]

class JobSchedule(BaseModel):
    scheduled_date: datetime
    installer_ids: Optional[List[str]] = None

class ItemAssignment(BaseModel):
    """Atribuição de itens específicos a instaladores"""
    item_indices: List[int]  # Índices dos itens/produtos a atribuir
    installer_ids: List[str]  # IDs dos instaladores

class CheckIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    job_id: str
    installer_id: str
    checkin_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    checkout_at: Optional[datetime] = None
    checkin_photo: Optional[str] = None  # Base64 encoded
    checkout_photo: Optional[str] = None  # Base64 encoded
    gps_lat: Optional[float] = None
    gps_long: Optional[float] = None
    gps_accuracy: Optional[float] = None
    checkout_gps_lat: Optional[float] = None
    checkout_gps_long: Optional[float] = None
    checkout_gps_accuracy: Optional[float] = None
    notes: Optional[str] = None
    duration_minutes: Optional[int] = None
    installed_m2: Optional[float] = None  # M² instalado
    # Campos de métricas de produtividade
    complexity_level: Optional[int] = None  # 1-5
    height_category: Optional[str] = None  # terreo, media, alta, muito_alta
    scenario_category: Optional[str] = None  # loja_rua, shopping, evento, fachada, outdoor, veiculo
    difficulty_description: Optional[str] = None  # Descrição da dificuldade
    productivity_m2_h: Optional[float] = None  # Produtividade calculada (m²/hora)
    status: str = "in_progress"  # in_progress, completed

class CheckInCreate(BaseModel):
    job_id: str
    gps_lat: Optional[float] = None
    gps_long: Optional[float] = None
    photo_base64: Optional[str] = None

class CheckOutUpdate(BaseModel):
    gps_lat: Optional[float] = None
    gps_long: Optional[float] = None
    photo_base64: Optional[str] = None
    notes: Optional[str] = None

# ============ PRODUCT FAMILIES & PRODUCTIVITY MODELS ============

class ProductFamily(BaseModel):
    """Família de produtos para categorização"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str  # Ex: "Adesivos", "Lonas", "ACM", etc.
    description: Optional[str] = None
    color: str = "#3B82F6"  # Cor para identificação visual
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProductFamilyCreate(BaseModel):
    name: str
    description: Optional[str] = None
    color: str = "#3B82F6"

class ProductInstalled(BaseModel):
    """Registro de cada produto instalado com métricas de produtividade"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    job_id: str
    checkin_id: Optional[str] = None
    product_name: str  # Nome do produto da Holdprint
    family_id: Optional[str] = None  # FK para ProductFamily
    family_name: Optional[str] = None  # Nome da família (desnormalizado para consultas rápidas)
    
    # Medidas
    width_m: Optional[float] = None
    height_m: Optional[float] = None
    quantity: int = 1
    area_m2: Optional[float] = None  # Calculado: width * height * quantity
    
    # Complexidade e contexto
    complexity_level: int = 1  # 1-5
    height_category: str = "terreo"  # terreo, media, alta, muito_alta
    scenario_category: str = "loja_rua"  # loja_rua, shopping, evento, fachada, etc.
    
    # Tempos
    estimated_time_min: Optional[int] = None
    actual_time_min: Optional[int] = None
    
    # Produtividade calculada
    productivity_m2_h: Optional[float] = None  # m²/hora
    
    # Metadados
    installers_count: int = 1
    installation_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    cause_notes: Optional[str] = None  # Causa de desvio, se houver
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProductInstalledCreate(BaseModel):
    job_id: str
    checkin_id: Optional[str] = None
    product_name: str
    family_id: Optional[str] = None
    width_m: Optional[float] = None
    height_m: Optional[float] = None
    quantity: int = 1
    complexity_level: int = 1
    height_category: str = "terreo"
    scenario_category: str = "loja_rua"
    estimated_time_min: Optional[int] = None
    actual_time_min: Optional[int] = None
    installers_count: int = 1
    cause_notes: Optional[str] = None

class ProductivityHistory(BaseModel):
    """Histórico consolidado de produtividade para benchmarks"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    family_id: str
    family_name: str
    complexity_level: int
    height_category: str
    scenario_category: str
    avg_productivity_m2_h: float
    avg_time_per_m2_min: float
    sample_count: int
    last_updated: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ============ UTILITY FUNCTIONS ============

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(days=ACCESS_TOKEN_EXPIRE_DAYS)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user_doc is None:
        raise credentials_exception
    return User(**user_doc)

async def require_role(user: User, allowed_roles: List[str]):
    if user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    return user

def compress_image_to_base64(image_data: bytes, max_size_kb: int = 500) -> str:
    """Compress image and return base64 string"""
    img = Image.open(BytesIO(image_data))
    
    # Convert to RGB if necessary
    if img.mode != 'RGB':
        img = img.convert('RGB')
    
    # Start with quality 85
    quality = 85
    output = BytesIO()
    
    while quality > 20:
        output = BytesIO()
        img.save(output, format='JPEG', quality=quality, optimize=True)
        size_kb = len(output.getvalue()) / 1024
        
        if size_kb <= max_size_kb:
            break
        quality -= 10
    
    return base64.b64encode(output.getvalue()).decode('utf-8')

async def fetch_holdprint_jobs(branch: str):
    """Fetch jobs from Holdprint API - últimos 7 dias, excluindo finalizados"""
    api_key = HOLDPRINT_API_KEY_POA if branch == "POA" else HOLDPRINT_API_KEY_SP
    
    if not api_key:
        raise HTTPException(status_code=500, detail=f"API key not configured for branch {branch}")
    
    headers = {"x-api-key": api_key}
    
    # Calcular datas: últimos 7 dias
    end_date = datetime.now(timezone.utc)
    start_date = end_date - timedelta(days=7)
    
    # Formatar datas no padrão YYYY-MM-DD
    start_date_str = start_date.strftime("%Y-%m-%d")
    end_date_str = end_date.strftime("%Y-%m-%d")
    
    # Montar URL com parâmetros de filtro
    params = {
        "page": 1,
        "pageSize": 100,
        "startDate": start_date_str,
        "endDate": end_date_str,
        "language": "pt-BR"
    }
    
    try:
        response = requests.get(HOLDPRINT_API_URL, headers=headers, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()
        
        # Holdprint returns {data: [...]} format
        jobs = []
        if isinstance(data, dict) and 'data' in data:
            jobs = data['data']
        elif isinstance(data, list):
            jobs = data
        
        # Filtrar jobs NÃO finalizados (isFinalized = false ou não existe)
        filtered_jobs = [job for job in jobs if not job.get('isFinalized', False)]
        
        logger.info(f"Holdprint {branch}: {len(jobs)} jobs encontrados, {len(filtered_jobs)} não finalizados (últimos 7 dias: {start_date_str} a {end_date_str})")
        
        return filtered_jobs
    except requests.RequestException as e:
        logger.error(f"Error fetching from Holdprint: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching from Holdprint: {str(e)}")

# ============ AUTH ROUTES ============

@api_router.post("/auth/register", response_model=User)
async def register(user_data: UserCreate, current_user: User = Depends(get_current_user)):
    """Admin creates new user"""
    await require_role(current_user, [UserRole.ADMIN])
    
    # Check if user exists
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create user
    user = User(
        email=user_data.email,
        name=user_data.name,
        role=user_data.role
    )
    
    user_dict = user.model_dump()
    user_dict['password_hash'] = get_password_hash(user_data.password)
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    
    await db.users.insert_one(user_dict)
    
    # If installer, create installer record
    if user_data.role == UserRole.INSTALLER:
        installer = Installer(
            user_id=user.id,
            full_name=user_data.name,
            branch="POA"  # Default, can be updated later
        )
        installer_dict = installer.model_dump()
        installer_dict['created_at'] = installer_dict['created_at'].isoformat()
        await db.installers.insert_one(installer_dict)
    
    return user

@api_router.post("/auth/login", response_model=Token)
async def login(credentials: UserLogin):
    # Find user
    user_doc = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Verify password
    if not verify_password(credentials.password, user_doc['password_hash']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Convert datetime
    if isinstance(user_doc['created_at'], str):
        user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
    
    user = User(**user_doc)
    
    # Create token
    access_token = create_access_token(data={"sub": user.id, "email": user.email, "role": user.role})
    
    return Token(access_token=access_token, token_type="bearer", user=user)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

# ============ USER MANAGEMENT ROUTES ============

@api_router.get("/users", response_model=List[User])
async def list_users(current_user: User = Depends(get_current_user)):
    await require_role(current_user, [UserRole.ADMIN])
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    
    for user in users:
        if isinstance(user['created_at'], str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
    
    return users

@api_router.put("/users/{user_id}", response_model=User)
async def update_user(user_id: str, user_data: dict, current_user: User = Depends(get_current_user)):
    await require_role(current_user, [UserRole.ADMIN])
    
    # Update user
    update_data = {k: v for k, v in user_data.items() if k not in ['id', 'created_at', 'password']}
    
    if 'password' in user_data:
        update_data['password_hash'] = get_password_hash(user_data['password'])
    
    result = await db.users.find_one_and_update(
        {"id": user_id},
        {"$set": update_data},
        return_document=True,
        projection={"_id": 0, "password_hash": 0}
    )
    
    if not result:
        raise HTTPException(status_code=404, detail="User not found")
    
    if isinstance(result['created_at'], str):
        result['created_at'] = datetime.fromisoformat(result['created_at'])
    
    return User(**result)

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    await require_role(current_user, [UserRole.ADMIN])
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted"}

# ============ HOLDPRINT & JOB ROUTES ============

@api_router.get("/holdprint/jobs/{branch}")
async def get_holdprint_jobs(branch: str, current_user: User = Depends(get_current_user)):
    """Fetch jobs from Holdprint API"""
    if branch not in ["POA", "SP"]:
        raise HTTPException(status_code=400, detail="Branch must be POA or SP")
    
    jobs = await fetch_holdprint_jobs(branch)
    return {"success": True, "jobs": jobs}

@api_router.post("/jobs", response_model=Job)
async def create_job(job_data: JobCreate, current_user: User = Depends(get_current_user)):
    """Import job from Holdprint to local database"""
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    # Check if job already exists
    existing = await db.jobs.find_one({"holdprint_job_id": job_data.holdprint_job_id})
    if existing:
        raise HTTPException(status_code=400, detail="Job already imported")
    
    # Fetch from Holdprint
    holdprint_jobs = await fetch_holdprint_jobs(job_data.branch)
    holdprint_job = next((j for j in holdprint_jobs if str(j.get('id')) == job_data.holdprint_job_id), None)
    
    if not holdprint_job:
        raise HTTPException(status_code=404, detail="Job not found in Holdprint")
    
    # Calcular área dos produtos
    products_with_area, total_area_m2, total_products, total_quantity = calculate_job_products_area(holdprint_job)
    
    # Create job
    job = Job(
        holdprint_job_id=job_data.holdprint_job_id,
        title=holdprint_job.get('title', 'Sem título'),
        client_name=holdprint_job.get('customerName', 'Cliente não informado'),
        client_address='',
        branch=job_data.branch,
        items=holdprint_job.get('production', {}).get('items', []),
        holdprint_data=holdprint_job,
        # Campos calculados
        area_m2=total_area_m2,
        products_with_area=products_with_area,
        total_products=total_products,
        total_quantity=total_quantity
    )
    
    job_dict = job.model_dump()
    job_dict['created_at'] = job_dict['created_at'].isoformat()
    if job_dict.get('scheduled_date'):
        job_dict['scheduled_date'] = job_dict['scheduled_date'].isoformat()
    
    await db.jobs.insert_one(job_dict)
    return job

@api_router.get("/jobs", response_model=List[Job])
async def list_jobs(current_user: User = Depends(get_current_user)):
    """List jobs based on user role"""
    query = {}
    
    # Installers only see their assigned jobs
    if current_user.role == UserRole.INSTALLER:
        installer = await db.installers.find_one({"user_id": current_user.id}, {"_id": 0})
        if installer:
            query["assigned_installers"] = installer['id']
        else:
            return []
    
    jobs = await db.jobs.find(query, {"_id": 0}).to_list(1000)
    
    for job in jobs:
        if isinstance(job['created_at'], str):
            job['created_at'] = datetime.fromisoformat(job['created_at'])
        if job.get('scheduled_date') and isinstance(job['scheduled_date'], str):
            job['scheduled_date'] = datetime.fromisoformat(job['scheduled_date'])
    
    return jobs

@api_router.get("/jobs/{job_id}", response_model=Job)
async def get_job(job_id: str, current_user: User = Depends(get_current_user)):
    job_doc = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job_doc:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if isinstance(job_doc['created_at'], str):
        job_doc['created_at'] = datetime.fromisoformat(job_doc['created_at'])
    if job_doc.get('scheduled_date') and isinstance(job_doc['scheduled_date'], str):
        job_doc['scheduled_date'] = datetime.fromisoformat(job_doc['scheduled_date'])
    
    return Job(**job_doc)

@api_router.put("/jobs/{job_id}/assign", response_model=Job)
async def assign_job(job_id: str, assign_data: JobAssign, current_user: User = Depends(get_current_user)):
    """Assign installers to a job"""
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    result = await db.jobs.find_one_and_update(
        {"id": job_id},
        {"$set": {"assigned_installers": assign_data.installer_ids}},
        return_document=True,
        projection={"_id": 0}
    )
    
    if not result:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if isinstance(result['created_at'], str):
        result['created_at'] = datetime.fromisoformat(result['created_at'])
    if result.get('scheduled_date') and isinstance(result['scheduled_date'], str):
        result['scheduled_date'] = datetime.fromisoformat(result['scheduled_date'])
    
    return Job(**result)

@api_router.put("/jobs/{job_id}/schedule", response_model=Job)
async def schedule_job(job_id: str, schedule_data: JobSchedule, current_user: User = Depends(get_current_user)):
    """Schedule a job"""
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    update_data = {"scheduled_date": schedule_data.scheduled_date.isoformat()}
    if schedule_data.installer_ids:
        update_data["assigned_installers"] = schedule_data.installer_ids
    
    result = await db.jobs.find_one_and_update(
        {"id": job_id},
        {"$set": update_data},
        return_document=True,
        projection={"_id": 0}
    )
    
    if not result:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if isinstance(result['created_at'], str):
        result['created_at'] = datetime.fromisoformat(result['created_at'])
    if result.get('scheduled_date') and isinstance(result['scheduled_date'], str):
        result['scheduled_date'] = datetime.fromisoformat(result['scheduled_date'])
    
    return Job(**result)

@api_router.put("/jobs/{job_id}", response_model=Job)
async def update_job(job_id: str, job_update: dict, current_user: User = Depends(get_current_user)):
    """Update job details (status, schedule, assignments, etc)"""
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    # Prepare update data
    update_data = {}
    
    # Handle allowed fields
    if "status" in job_update:
        update_data["status"] = job_update["status"]
    
    if "scheduled_date" in job_update:
        # Convert to datetime if string
        if isinstance(job_update["scheduled_date"], str):
            update_data["scheduled_date"] = job_update["scheduled_date"]
        else:
            update_data["scheduled_date"] = job_update["scheduled_date"].isoformat()
    
    if "assigned_installers" in job_update:
        update_data["assigned_installers"] = job_update["assigned_installers"]
    
    if "client_name" in job_update:
        update_data["client_name"] = job_update["client_name"]
    
    if "client_address" in job_update:
        update_data["client_address"] = job_update["client_address"]
    
    if "title" in job_update:
        update_data["title"] = job_update["title"]
    
    if "area_m2" in job_update:
        update_data["area_m2"] = job_update["area_m2"]
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    
    result = await db.jobs.find_one_and_update(
        {"id": job_id},
        {"$set": update_data},
        return_document=True,
        projection={"_id": 0}
    )
    
    if not result:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if isinstance(result['created_at'], str):
        result['created_at'] = datetime.fromisoformat(result['created_at'])
    if result.get('scheduled_date') and isinstance(result['scheduled_date'], str):
        result['scheduled_date'] = datetime.fromisoformat(result['scheduled_date'])
    
    return Job(**result)

@api_router.post("/jobs/{job_id}/assign-items")
async def assign_items_to_installers(job_id: str, assignment: ItemAssignment, current_user: User = Depends(get_current_user)):
    """
    Atribui itens específicos do job a instaladores.
    Permite selecionar múltiplos itens e atribuir a um ou mais instaladores.
    """
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    # Buscar job
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Buscar instaladores
    installers = await db.installers.find({"id": {"$in": assignment.installer_ids}}, {"_id": 0}).to_list(100)
    installer_map = {i["id"]: i for i in installers}
    
    if len(installers) != len(assignment.installer_ids):
        raise HTTPException(status_code=400, detail="One or more installers not found")
    
    # Validar índices dos itens
    products = job.get("products_with_area", [])
    if not products:
        # Se não tiver products_with_area, usar holdprint_data.products
        products = job.get("holdprint_data", {}).get("products", [])
    
    for idx in assignment.item_indices:
        if idx < 0 or idx >= len(products):
            raise HTTPException(status_code=400, detail=f"Invalid item index: {idx}")
    
    # Criar atribuições
    current_assignments = job.get("item_assignments", [])
    now = datetime.now(timezone.utc).isoformat()
    
    new_assignments = []
    total_m2_assigned = 0
    
    for item_idx in assignment.item_indices:
        product = products[item_idx] if item_idx < len(products) else None
        item_area = product.get("total_area_m2") if product else 0
        item_area = item_area if item_area is not None else 0
        
        for installer_id in assignment.installer_ids:
            installer = installer_map.get(installer_id)
            
            # Remover atribuição anterior do mesmo item (se existir)
            current_assignments = [a for a in current_assignments 
                                  if not (a.get("item_index") == item_idx and a.get("installer_id") == installer_id)]
            
            # Calcular m² por instalador (dividir igualmente se múltiplos instaladores)
            m2_per_installer = round(item_area / len(assignment.installer_ids), 2) if item_area and item_area > 0 else 0
            
            new_assignment = {
                "item_index": item_idx,
                "item_name": product.get("name", f"Item {item_idx}") if product else f"Item {item_idx}",
                "installer_id": installer_id,
                "installer_name": installer.get("full_name", ""),
                "assigned_at": now,
                "item_area_m2": item_area,
                "assigned_m2": m2_per_installer,
                "status": "pending"  # pending, in_progress, completed
            }
            new_assignments.append(new_assignment)
            total_m2_assigned += m2_per_installer
    
    # Combinar atribuições
    all_assignments = current_assignments + new_assignments
    
    # Atualizar assigned_installers do job (lista única de IDs)
    all_installer_ids = list(set([a["installer_id"] for a in all_assignments]))
    
    # Atualizar job
    await db.jobs.update_one(
        {"id": job_id},
        {"$set": {
            "item_assignments": all_assignments,
            "assigned_installers": all_installer_ids
        }}
    )
    
    return {
        "message": f"{len(new_assignments)} atribuições criadas",
        "total_m2_assigned": total_m2_assigned,
        "assignments": new_assignments
    }

@api_router.get("/jobs/{job_id}/assignments")
async def get_job_assignments(job_id: str, current_user: User = Depends(get_current_user)):
    """
    Retorna as atribuições de itens do job, agrupadas por instalador.
    """
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER, UserRole.INSTALLER])
    
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    assignments = job.get("item_assignments", [])
    products = job.get("products_with_area", []) or job.get("holdprint_data", {}).get("products", [])
    
    # Agrupar por instalador
    by_installer = {}
    for assignment in assignments:
        installer_id = assignment.get("installer_id")
        if installer_id not in by_installer:
            by_installer[installer_id] = {
                "installer_id": installer_id,
                "installer_name": assignment.get("installer_name"),
                "items": [],
                "total_m2": 0
            }
        
        by_installer[installer_id]["items"].append(assignment)
        by_installer[installer_id]["total_m2"] += assignment.get("assigned_m2", 0)
    
    # Agrupar por item
    by_item = {}
    for assignment in assignments:
        item_idx = assignment.get("item_index")
        if item_idx not in by_item:
            product = products[item_idx] if item_idx < len(products) else {}
            item_area = product.get("total_area_m2", 0) or 0
            by_item[item_idx] = {
                "item_index": item_idx,
                "item_name": product.get("name", f"Item {item_idx}"),
                "item_area_m2": item_area,
                "installers": []
            }
        
        by_item[item_idx]["installers"].append({
            "installer_id": assignment.get("installer_id"),
            "installer_name": assignment.get("installer_name"),
            "assigned_m2": assignment.get("assigned_m2"),
            "status": assignment.get("status")
        })
    
    return {
        "job_id": job_id,
        "job_title": job.get("title"),
        "total_area_m2": job.get("area_m2", 0),
        "by_installer": list(by_installer.values()),
        "by_item": list(by_item.values()),
        "all_assignments": assignments
    }

@api_router.put("/jobs/{job_id}/assignments/{item_index}/status")
async def update_assignment_status(job_id: str, item_index: int, status_update: dict, current_user: User = Depends(get_current_user)):
    """
    Atualiza o status de uma atribuição de item (instalador reportando progresso).
    """
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    new_status = status_update.get("status")
    installed_m2 = status_update.get("installed_m2")
    
    if new_status not in ["pending", "in_progress", "completed"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    assignments = job.get("item_assignments", [])
    updated = False
    
    for assignment in assignments:
        if assignment.get("item_index") == item_index:
            # Se for instalador, só pode atualizar sua própria atribuição
            if current_user.role == UserRole.INSTALLER:
                installer = await db.installers.find_one({"user_id": current_user.id}, {"_id": 0})
                if not installer or installer.get("id") != assignment.get("installer_id"):
                    continue
            
            assignment["status"] = new_status
            if installed_m2 is not None:
                assignment["installed_m2"] = installed_m2
            if new_status == "completed":
                assignment["completed_at"] = datetime.now(timezone.utc).isoformat()
            updated = True
    
    if not updated:
        raise HTTPException(status_code=404, detail="Assignment not found or unauthorized")
    
    await db.jobs.update_one(
        {"id": job_id},
        {"$set": {"item_assignments": assignments}}
    )
    
    return {"message": "Assignment status updated", "assignments": assignments}

# ============ CHECK-IN/OUT ROUTES ============

# Create uploads directory if it doesn't exist
UPLOAD_DIR = Path("/app/uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

@api_router.post("/checkins", response_model=CheckIn)
async def create_checkin(
    job_id: str = Form(...),
    photo_base64: str = Form(...),
    gps_lat: float = Form(...),
    gps_long: float = Form(...),
    gps_accuracy: Optional[float] = Form(None),
    current_user: User = Depends(get_current_user)
):
    """Create check-in for a job with photo in Base64 and GPS coordinates"""
    # Get installer
    installer = await db.installers.find_one({"user_id": current_user.id}, {"_id": 0})
    if not installer:
        raise HTTPException(status_code=400, detail="User is not an installer")
    
    # Check if job exists
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Check for existing open checkin
    existing = await db.checkins.find_one({
        "job_id": job_id,
        "installer_id": installer['id'],
        "status": "in_progress"
    })
    if existing:
        raise HTTPException(status_code=400, detail="Already checked in")
    
    # Create checkin with Base64 photo and GPS
    checkin_id = str(uuid.uuid4())
    checkin = CheckIn(
        id=checkin_id,
        job_id=job_id,
        installer_id=installer['id'],
        checkin_photo=photo_base64,
        gps_lat=gps_lat,
        gps_long=gps_long,
        gps_accuracy=gps_accuracy
    )
    
    checkin_dict = checkin.model_dump()
    checkin_dict['checkin_at'] = checkin_dict['checkin_at'].isoformat()
    if checkin_dict.get('checkout_at'):
        checkin_dict['checkout_at'] = checkin_dict['checkout_at'].isoformat()
    
    await db.checkins.insert_one(checkin_dict)
    
    # Update job status
    await db.jobs.update_one(
        {"id": job_id},
        {"$set": {"status": "in_progress"}}
    )
    
    return checkin

@api_router.put("/checkins/{checkin_id}/checkout", response_model=CheckIn)
async def checkout(
    checkin_id: str,
    photo_base64: str = Form(...),
    gps_lat: float = Form(...),
    gps_long: float = Form(...),
    gps_accuracy: Optional[float] = Form(None),
    installed_m2: Optional[float] = Form(None),
    complexity_level: Optional[int] = Form(None),
    height_category: Optional[str] = Form(None),
    scenario_category: Optional[str] = Form(None),
    difficulty_description: Optional[str] = Form(None),
    notes: str = Form(""),
    current_user: User = Depends(get_current_user)
):
    """Check out from a job with photo in Base64, GPS coordinates and productivity metrics"""
    checkin_doc = await db.checkins.find_one({"id": checkin_id}, {"_id": 0})
    if not checkin_doc:
        raise HTTPException(status_code=404, detail="Check-in not found")
    
    if checkin_doc['status'] == "completed":
        raise HTTPException(status_code=400, detail="Already checked out")
    
    # Calculate duration
    checkout_at = datetime.now(timezone.utc)
    checkin_at = datetime.fromisoformat(checkin_doc['checkin_at']) if isinstance(checkin_doc['checkin_at'], str) else checkin_doc['checkin_at']
    duration_minutes = int((checkout_at - checkin_at).total_seconds() / 60)
    
    # Calculate productivity if m2 and duration available
    productivity_m2_h = None
    if installed_m2 and installed_m2 > 0 and duration_minutes > 0:
        hours = duration_minutes / 60
        productivity_m2_h = round(installed_m2 / hours, 2)
    
    # Update checkin with Base64 photo, GPS and metrics
    update_data = {
        "checkout_at": checkout_at.isoformat(),
        "checkout_photo": photo_base64,
        "checkout_gps_lat": gps_lat,
        "checkout_gps_long": gps_long,
        "checkout_gps_accuracy": gps_accuracy,
        "installed_m2": installed_m2,
        "complexity_level": complexity_level,
        "height_category": height_category,
        "scenario_category": scenario_category,
        "difficulty_description": difficulty_description,
        "productivity_m2_h": productivity_m2_h,
        "notes": notes,
        "duration_minutes": duration_minutes,
        "status": "completed"
    }
    
    result = await db.checkins.find_one_and_update(
        {"id": checkin_id},
        {"$set": update_data},
        return_document=True,
        projection={"_id": 0}
    )
    
    # Check if all checkins for this job are completed
    job_checkins = await db.checkins.find({"job_id": checkin_doc['job_id']}, {"_id": 0}).to_list(1000)
    all_completed = all(c['status'] == "completed" for c in job_checkins)
    
    if all_completed:
        await db.jobs.update_one(
            {"id": checkin_doc['job_id']},
            {"$set": {"status": "completed"}}
        )
    
    if isinstance(result['checkin_at'], str):
        result['checkin_at'] = datetime.fromisoformat(result['checkin_at'])
    if result.get('checkout_at') and isinstance(result['checkout_at'], str):
        result['checkout_at'] = datetime.fromisoformat(result['checkout_at'])
    
    return CheckIn(**result)

@api_router.get("/checkins", response_model=List[CheckIn])
async def list_checkins(job_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """List check-ins"""
    query = {}
    
    if job_id:
        query["job_id"] = job_id
    
    # Installers only see their own checkins
    if current_user.role == UserRole.INSTALLER:
        installer = await db.installers.find_one({"user_id": current_user.id}, {"_id": 0})
        if installer:
            query["installer_id"] = installer['id']
        else:
            return []
    
    checkins = await db.checkins.find(query, {"_id": 0}).to_list(1000)
    
    for checkin in checkins:
        if isinstance(checkin['checkin_at'], str):
            checkin['checkin_at'] = datetime.fromisoformat(checkin['checkin_at'])
        if checkin.get('checkout_at') and isinstance(checkin['checkout_at'], str):
            checkin['checkout_at'] = datetime.fromisoformat(checkin['checkout_at'])
    
    return checkins

@api_router.get("/checkins/{checkin_id}/details")
async def get_checkin_details(
    checkin_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get check-in details with photos and GPS data for managers/admins"""
    # Only admin and managers can view detailed checkin data
    if current_user.role not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    
    checkin = await db.checkins.find_one({"id": checkin_id}, {"_id": 0})
    if not checkin:
        raise HTTPException(status_code=404, detail="Check-in not found")
    
    # Get installer info
    installer = await db.installers.find_one({"id": checkin['installer_id']}, {"_id": 0})
    
    # Get job info
    job = await db.jobs.find_one({"id": checkin['job_id']}, {"_id": 0})
    
    return {
        "checkin": checkin,
        "installer": installer,
        "job": job
    }

# ============ INSTALLER ROUTES ============

@api_router.get("/installers", response_model=List[Installer])
async def list_installers(current_user: User = Depends(get_current_user)):
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    installers = await db.installers.find({}, {"_id": 0}).to_list(1000)
    
    for installer in installers:
        if isinstance(installer['created_at'], str):
            installer['created_at'] = datetime.fromisoformat(installer['created_at'])
    
    return installers

@api_router.put("/installers/{installer_id}", response_model=Installer)
async def update_installer(installer_id: str, installer_data: dict, current_user: User = Depends(get_current_user)):
    await require_role(current_user, [UserRole.ADMIN])
    
    update_data = {k: v for k, v in installer_data.items() if k not in ['id', 'user_id', 'created_at']}
    
    result = await db.installers.find_one_and_update(
        {"id": installer_id},
        {"$set": update_data},
        return_document=True,
        projection={"_id": 0}
    )
    
    if not result:
        raise HTTPException(status_code=404, detail="Installer not found")
    
    if isinstance(result['created_at'], str):
        result['created_at'] = datetime.fromisoformat(result['created_at'])
    
    return Installer(**result)

# ============ METRICS ROUTES ============

# ============ PRODUCT FAMILIES ENDPOINTS ============

@api_router.get("/product-families")
async def get_product_families(current_user: User = Depends(get_current_user)):
    """List all product families"""
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    families = await db.product_families.find({}, {"_id": 0}).to_list(1000)
    return families

@api_router.post("/product-families")
async def create_product_family(family: ProductFamilyCreate, current_user: User = Depends(get_current_user)):
    """Create a new product family"""
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    new_family = ProductFamily(**family.model_dump())
    await db.product_families.insert_one(new_family.model_dump())
    return new_family.model_dump()

@api_router.put("/product-families/{family_id}")
async def update_product_family(family_id: str, family: ProductFamilyCreate, current_user: User = Depends(get_current_user)):
    """Update a product family"""
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    result = await db.product_families.update_one(
        {"id": family_id},
        {"$set": family.model_dump()}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Family not found")
    
    updated = await db.product_families.find_one({"id": family_id}, {"_id": 0})
    return updated

@api_router.delete("/product-families/{family_id}")
async def delete_product_family(family_id: str, current_user: User = Depends(get_current_user)):
    """Delete a product family"""
    await require_role(current_user, [UserRole.ADMIN])
    
    result = await db.product_families.delete_one({"id": family_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Family not found")
    return {"message": "Family deleted"}

@api_router.post("/product-families/seed")
async def seed_product_families(current_user: User = Depends(get_current_user)):
    """Seed initial product families from Holdprint catalog"""
    await require_role(current_user, [UserRole.ADMIN])
    
    # Famílias padrão baseadas no catálogo Holdprint
    default_families = [
        {"name": "Adesivos", "description": "Adesivos impressos, coloridos, jateados, etc.", "color": "#EF4444"},
        {"name": "Lonas e Banners", "description": "Lonas frontlight, backlight, banners", "color": "#F97316"},
        {"name": "Chapas e Placas", "description": "ACM, acrílico, PVC, PS, MDF com ou sem impressão", "color": "#EAB308"},
        {"name": "Estruturas Metálicas", "description": "Estruturas com lona, ACM ou chapa galvanizada", "color": "#22C55E"},
        {"name": "Tecidos", "description": "Bandeiras, faixas, wind banners em tecido", "color": "#14B8A6"},
        {"name": "Letras Caixa", "description": "Letras planas, em relevo, iluminadas", "color": "#3B82F6"},
        {"name": "Totens", "description": "Totens em diversos materiais e formatos", "color": "#8B5CF6"},
        {"name": "Envelopamento", "description": "Envelopamento de veículos", "color": "#EC4899"},
        {"name": "Painéis Luminosos", "description": "Backlight, painéis com iluminação", "color": "#F59E0B"},
        {"name": "Serviços", "description": "Instalação, entrega, montagem, pintura", "color": "#6B7280"},
        {"name": "Materiais Promocionais", "description": "Cartazes, flyers, folders, panfletos", "color": "#84CC16"},
        {"name": "Produtos Terceirizados", "description": "Produtos de terceiros", "color": "#A855F7"},
    ]
    
    inserted = 0
    for family_data in default_families:
        existing = await db.product_families.find_one({"name": family_data["name"]})
        if not existing:
            new_family = ProductFamily(**family_data)
            await db.product_families.insert_one(new_family.model_dump())
            inserted += 1
    
    return {"message": f"{inserted} families created", "total": len(default_families)}

# ============ PRODUCTS INSTALLED ENDPOINTS ============

@api_router.get("/products-installed")
async def get_products_installed(
    job_id: Optional[str] = None,
    family_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """List installed products with optional filters"""
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    query = {}
    if job_id:
        query["job_id"] = job_id
    if family_id:
        query["family_id"] = family_id
    
    products = await db.products_installed.find(query, {"_id": 0}).to_list(1000)
    return products

@api_router.post("/products-installed")
async def create_product_installed(product: ProductInstalledCreate, current_user: User = Depends(get_current_user)):
    """Register a new installed product with productivity metrics"""
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER, UserRole.INSTALLER])
    
    # Calculate area
    area_m2 = None
    if product.width_m and product.height_m:
        area_m2 = product.width_m * product.height_m * product.quantity
    
    # Calculate productivity (m²/hour)
    productivity_m2_h = None
    if area_m2 and product.actual_time_min and product.actual_time_min > 0:
        hours = product.actual_time_min / 60
        productivity_m2_h = round(area_m2 / hours, 2)
    
    # Get family name if family_id provided
    family_name = None
    if product.family_id:
        family = await db.product_families.find_one({"id": product.family_id}, {"_id": 0})
        if family:
            family_name = family.get("name")
    
    new_product = ProductInstalled(
        **product.model_dump(),
        area_m2=area_m2,
        productivity_m2_h=productivity_m2_h,
        family_name=family_name
    )
    
    await db.products_installed.insert_one(new_product.model_dump())
    
    # Update productivity history
    await update_productivity_history(new_product)
    
    return new_product.model_dump()

async def update_productivity_history(product: ProductInstalled):
    """Update the productivity history based on new data"""
    if not product.family_id or not product.productivity_m2_h:
        return
    
    key = {
        "family_id": product.family_id,
        "complexity_level": product.complexity_level,
        "height_category": product.height_category,
        "scenario_category": product.scenario_category
    }
    
    existing = await db.productivity_history.find_one(key, {"_id": 0})
    
    if existing:
        # Calculate new average
        new_count = existing["sample_count"] + 1
        new_avg_prod = ((existing["avg_productivity_m2_h"] * existing["sample_count"]) + product.productivity_m2_h) / new_count
        
        # Calculate avg time per m2
        new_avg_time = 60 / new_avg_prod if new_avg_prod > 0 else 0
        
        await db.productivity_history.update_one(
            key,
            {
                "$set": {
                    "avg_productivity_m2_h": round(new_avg_prod, 2),
                    "avg_time_per_m2_min": round(new_avg_time, 2),
                    "sample_count": new_count,
                    "last_updated": datetime.now(timezone.utc).isoformat()
                }
            }
        )
    else:
        # Create new record
        avg_time = 60 / product.productivity_m2_h if product.productivity_m2_h > 0 else 0
        new_history = ProductivityHistory(
            family_id=product.family_id,
            family_name=product.family_name or "",
            complexity_level=product.complexity_level,
            height_category=product.height_category,
            scenario_category=product.scenario_category,
            avg_productivity_m2_h=product.productivity_m2_h,
            avg_time_per_m2_min=round(avg_time, 2),
            sample_count=1
        )
        await db.productivity_history.insert_one(new_history.model_dump())

@api_router.get("/productivity-history")
async def get_productivity_history(
    family_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get productivity benchmarks"""
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    query = {}
    if family_id:
        query["family_id"] = family_id
    
    history = await db.productivity_history.find(query, {"_id": 0}).to_list(1000)
    return history

@api_router.get("/productivity-metrics")
async def get_productivity_metrics(current_user: User = Depends(get_current_user)):
    """Get comprehensive productivity metrics"""
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    # Get all product families
    families = await db.product_families.find({}, {"_id": 0}).to_list(100)
    
    # Get all products installed
    products = await db.products_installed.find({}, {"_id": 0}).to_list(10000)
    
    # Get productivity history
    history = await db.productivity_history.find({}, {"_id": 0}).to_list(1000)
    
    # Calculate metrics by family
    family_metrics = {}
    for family in families:
        family_products = [p for p in products if p.get("family_id") == family["id"]]
        
        total_area = sum(p.get("area_m2", 0) or 0 for p in family_products)
        total_time = sum(p.get("actual_time_min", 0) or 0 for p in family_products)
        total_products = len(family_products)
        
        avg_productivity = 0
        if total_time > 0:
            avg_productivity = round((total_area / (total_time / 60)), 2) if total_area > 0 else 0
        
        family_metrics[family["name"]] = {
            "family_id": family["id"],
            "color": family.get("color", "#3B82F6"),
            "total_products": total_products,
            "total_area_m2": round(total_area, 2),
            "total_time_hours": round(total_time / 60, 2),
            "avg_productivity_m2_h": avg_productivity,
            "avg_time_per_m2_min": round(60 / avg_productivity, 2) if avg_productivity > 0 else 0
        }
    
    # Calculate overall metrics
    total_area_all = sum(p.get("area_m2", 0) or 0 for p in products)
    total_time_all = sum(p.get("actual_time_min", 0) or 0 for p in products)
    overall_productivity = round((total_area_all / (total_time_all / 60)), 2) if total_time_all > 0 and total_area_all > 0 else 0
    
    # Metrics by complexity
    complexity_metrics = {}
    for level in [1, 2, 3, 4, 5]:
        level_products = [p for p in products if p.get("complexity_level") == level]
        total_area = sum(p.get("area_m2", 0) or 0 for p in level_products)
        total_time = sum(p.get("actual_time_min", 0) or 0 for p in level_products)
        
        complexity_metrics[f"level_{level}"] = {
            "total_products": len(level_products),
            "total_area_m2": round(total_area, 2),
            "avg_productivity_m2_h": round((total_area / (total_time / 60)), 2) if total_time > 0 and total_area > 0 else 0
        }
    
    # Metrics by height category
    height_metrics = {}
    for category in ["terreo", "media", "alta", "muito_alta"]:
        cat_products = [p for p in products if p.get("height_category") == category]
        total_area = sum(p.get("area_m2", 0) or 0 for p in cat_products)
        total_time = sum(p.get("actual_time_min", 0) or 0 for p in cat_products)
        
        height_metrics[category] = {
            "total_products": len(cat_products),
            "total_area_m2": round(total_area, 2),
            "avg_productivity_m2_h": round((total_area / (total_time / 60)), 2) if total_time > 0 and total_area > 0 else 0
        }
    
    # Metrics by scenario
    scenario_metrics = {}
    for scenario in ["loja_rua", "shopping", "evento", "fachada", "outdoor", "veiculo"]:
        scen_products = [p for p in products if p.get("scenario_category") == scenario]
        total_area = sum(p.get("area_m2", 0) or 0 for p in scen_products)
        total_time = sum(p.get("actual_time_min", 0) or 0 for p in scen_products)
        
        scenario_metrics[scenario] = {
            "total_products": len(scen_products),
            "total_area_m2": round(total_area, 2),
            "avg_productivity_m2_h": round((total_area / (total_time / 60)), 2) if total_time > 0 and total_area > 0 else 0
        }
    
    return {
        "overall": {
            "total_products": len(products),
            "total_area_m2": round(total_area_all, 2),
            "total_time_hours": round(total_time_all / 60, 2),
            "avg_productivity_m2_h": overall_productivity
        },
        "by_family": family_metrics,
        "by_complexity": complexity_metrics,
        "by_height": height_metrics,
        "by_scenario": scenario_metrics,
        "benchmarks": history
    }

@api_router.get("/reports/by-family")
async def get_report_by_family(current_user: User = Depends(get_current_user)):
    """
    Relatório completo por família de produtos.
    Analisa todos os jobs importados e classifica seus produtos por família.
    """
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    # Buscar todos os jobs
    jobs = await db.jobs.find({}, {"_id": 0}).to_list(10000)
    
    # Buscar famílias cadastradas
    families = await db.product_families.find({}, {"_id": 0}).to_list(100)
    family_map = {f["name"]: f for f in families}
    
    # Estrutura para agrupar dados por família
    family_report = {}
    all_products = []
    unclassified_products = []
    
    for job in jobs:
        holdprint_data = job.get("holdprint_data", {})
        products = holdprint_data.get("products", [])
        production_items = holdprint_data.get("production", {}).get("items", [])
        
        # Processar produtos do job
        for product in products:
            product_name = product.get("name", "")
            quantity = product.get("quantity", 1)
            
            # Extrair medidas da descrição
            description = product.get("description", "")
            width_m = None
            height_m = None
            
            # Parse de medidas da descrição HTML
            import re
            width_match = re.search(r'Largura:\s*<span[^>]*>([0-9.,]+)\s*m', description, re.IGNORECASE)
            height_match = re.search(r'Altura:\s*<span[^>]*>([0-9.,]+)\s*m', description, re.IGNORECASE)
            
            if width_match:
                width_m = float(width_match.group(1).replace(',', '.'))
            if height_match:
                height_m = float(height_match.group(1).replace(',', '.'))
            
            # Calcular área
            area_m2 = None
            if width_m and height_m:
                area_m2 = round(width_m * height_m * quantity, 2)
            
            # Classificar produto em família
            family_name, confidence = classify_product_to_family(product_name)
            
            product_data = {
                "job_id": job.get("id"),
                "job_title": job.get("title"),
                "job_code": holdprint_data.get("code"),
                "client_name": holdprint_data.get("customerName", job.get("client_name")),
                "product_name": product_name,
                "family_name": family_name,
                "confidence": confidence,
                "quantity": quantity,
                "width_m": width_m,
                "height_m": height_m,
                "area_m2": area_m2,
                "unit_price": product.get("unitPrice", 0),
                "total_value": product.get("totalValue", 0),
                "branch": job.get("branch")
            }
            
            all_products.append(product_data)
            
            # Agrupar por família
            if family_name not in family_report:
                family_info = family_map.get(family_name, {})
                family_report[family_name] = {
                    "family_name": family_name,
                    "color": family_info.get("color", "#6B7280"),
                    "total_jobs": set(),
                    "total_products": 0,
                    "total_quantity": 0,
                    "total_area_m2": 0,
                    "total_value": 0,
                    "products": []
                }
            
            family_report[family_name]["total_jobs"].add(job.get("id"))
            family_report[family_name]["total_products"] += 1
            family_report[family_name]["total_quantity"] += quantity
            if area_m2:
                family_report[family_name]["total_area_m2"] += area_m2
            family_report[family_name]["total_value"] += product.get("totalValue", 0)
            family_report[family_name]["products"].append(product_data)
            
            # Rastrear produtos não classificados com alta confiança
            if confidence < 50:
                unclassified_products.append(product_data)
        
        # Processar itens de produção também
        for item in production_items:
            item_name = item.get("name", "")
            item_quantity = item.get("quantity", 1)
            
            family_name, confidence = classify_product_to_family(item_name)
            
            if family_name not in family_report:
                family_info = family_map.get(family_name, {})
                family_report[family_name] = {
                    "family_name": family_name,
                    "color": family_info.get("color", "#6B7280"),
                    "total_jobs": set(),
                    "total_products": 0,
                    "total_quantity": 0,
                    "total_area_m2": 0,
                    "total_value": 0,
                    "products": []
                }
            
            family_report[family_name]["total_jobs"].add(job.get("id"))
            family_report[family_name]["total_quantity"] += item_quantity
    
    # Converter sets para contagem
    for family_name in family_report:
        family_report[family_name]["total_jobs"] = len(family_report[family_name]["total_jobs"])
        family_report[family_name]["total_area_m2"] = round(family_report[family_name]["total_area_m2"], 2)
        family_report[family_name]["total_value"] = round(family_report[family_name]["total_value"], 2)
        # Limitar lista de produtos para não sobrecarregar resposta
        family_report[family_name]["products"] = family_report[family_name]["products"][:50]
    
    # Ordenar por quantidade total
    sorted_families = sorted(
        family_report.values(),
        key=lambda x: x["total_quantity"],
        reverse=True
    )
    
    # Estatísticas gerais
    total_area = sum(f["total_area_m2"] for f in sorted_families)
    total_value = sum(f["total_value"] for f in sorted_families)
    total_products = sum(f["total_products"] for f in sorted_families)
    
    return {
        "summary": {
            "total_jobs": len(jobs),
            "total_products": total_products,
            "total_area_m2": round(total_area, 2),
            "total_value": round(total_value, 2),
            "families_count": len(sorted_families),
            "unclassified_count": len(unclassified_products)
        },
        "by_family": sorted_families,
        "unclassified": unclassified_products[:20],  # Primeiros 20 não classificados
        "all_products": all_products[:100]  # Primeiros 100 produtos para análise
    }

@api_router.post("/jobs/{job_id}/classify-products")
async def classify_job_products(job_id: str, current_user: User = Depends(get_current_user)):
    """
    Classifica os produtos de um job específico por família.
    Retorna a análise detalhada para esse job.
    """
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    holdprint_data = job.get("holdprint_data", {})
    products = holdprint_data.get("products", [])
    
    classified_products = []
    
    for product in products:
        product_name = product.get("name", "")
        family_name, confidence = classify_product_to_family(product_name)
        
        # Extrair medidas
        description = product.get("description", "")
        import re
        width_match = re.search(r'Largura:\s*<span[^>]*>([0-9.,]+)\s*m', description, re.IGNORECASE)
        height_match = re.search(r'Altura:\s*<span[^>]*>([0-9.,]+)\s*m', description, re.IGNORECASE)
        
        width_m = float(width_match.group(1).replace(',', '.')) if width_match else None
        height_m = float(height_match.group(1).replace(',', '.')) if height_match else None
        
        area_m2 = round(width_m * height_m * product.get("quantity", 1), 2) if width_m and height_m else None
        
        classified_products.append({
            "product_name": product_name,
            "family_name": family_name,
            "confidence": confidence,
            "quantity": product.get("quantity", 1),
            "width_m": width_m,
            "height_m": height_m,
            "area_m2": area_m2,
            "unit_price": product.get("unitPrice", 0),
            "total_value": product.get("totalValue", 0)
        })
    
    # Agrupar por família
    family_summary = {}
    for p in classified_products:
        fname = p["family_name"]
        if fname not in family_summary:
            family_summary[fname] = {
                "count": 0,
                "total_area_m2": 0,
                "total_value": 0
            }
        family_summary[fname]["count"] += 1
        if p["area_m2"]:
            family_summary[fname]["total_area_m2"] += p["area_m2"]
        family_summary[fname]["total_value"] += p["total_value"]
    
    return {
        "job_id": job_id,
        "job_title": job.get("title"),
        "client": holdprint_data.get("customerName"),
        "products": classified_products,
        "family_summary": family_summary
    }

@api_router.post("/jobs/recalculate-areas")
async def recalculate_job_areas(current_user: User = Depends(get_current_user)):
    """
    Recalcula a área de todos os jobs existentes.
    Útil para atualizar jobs importados antes da implementação do cálculo automático.
    """
    await require_role(current_user, [UserRole.ADMIN])
    
    jobs = await db.jobs.find({}, {"_id": 0}).to_list(10000)
    updated_count = 0
    
    for job in jobs:
        holdprint_data = job.get("holdprint_data", {})
        
        if holdprint_data:
            products_with_area, total_area_m2, total_products, total_quantity = calculate_job_products_area(holdprint_data)
            
            await db.jobs.update_one(
                {"id": job["id"]},
                {"$set": {
                    "area_m2": total_area_m2,
                    "products_with_area": products_with_area,
                    "total_products": total_products,
                    "total_quantity": total_quantity
                }}
            )
            updated_count += 1
    
    return {"message": f"{updated_count} jobs atualizados com áreas calculadas"}

@api_router.get("/reports/by-installer")
async def get_report_by_installer(current_user: User = Depends(get_current_user)):
    """
    Relatório de produtividade por instalador.
    Mostra m² instalados, jobs realizados, tempo médio por m².
    """
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    # Buscar dados
    installers = await db.installers.find({}, {"_id": 0}).to_list(1000)
    checkins = await db.checkins.find({}, {"_id": 0}).to_list(10000)
    jobs = await db.jobs.find({}, {"_id": 0}).to_list(10000)
    
    # Mapear jobs por ID
    jobs_map = {job["id"]: job for job in jobs}
    
    # Processar dados por instalador
    installer_report = []
    
    for installer in installers:
        installer_id = installer["id"]
        
        # Checkins deste instalador
        installer_checkins = [c for c in checkins if c.get("installer_id") == installer_id]
        
        # Calcular métricas
        total_checkins = len(installer_checkins)
        completed_checkins = [c for c in installer_checkins if c.get("status") == "completed"]
        total_duration_min = sum(c.get("duration_minutes", 0) or 0 for c in completed_checkins)
        total_m2_reported = sum(c.get("installed_m2", 0) or 0 for c in completed_checkins)
        
        # Jobs únicos trabalhados
        job_ids = set(c.get("job_id") for c in installer_checkins if c.get("job_id"))
        jobs_worked = len(job_ids)
        
        # Área total dos jobs trabalhados (estimada)
        total_job_area_m2 = 0
        jobs_details = []
        for job_id in job_ids:
            job = jobs_map.get(job_id)
            if job:
                job_area = job.get("area_m2", 0) or 0
                total_job_area_m2 += job_area
                
                # Checkins deste instalador neste job
                job_checkins = [c for c in installer_checkins if c.get("job_id") == job_id]
                job_duration = sum(c.get("duration_minutes", 0) or 0 for c in job_checkins if c.get("status") == "completed")
                job_m2_reported = sum(c.get("installed_m2", 0) or 0 for c in job_checkins if c.get("status") == "completed")
                
                jobs_details.append({
                    "job_id": job_id,
                    "job_title": job.get("title"),
                    "client": job.get("client_name") or job.get("holdprint_data", {}).get("customerName"),
                    "job_area_m2": job_area,
                    "duration_min": job_duration,
                    "m2_reported": job_m2_reported,
                    "status": job.get("status"),
                    "checkins_count": len(job_checkins)
                })
        
        # Produtividade (m²/hora)
        productivity_m2_h = 0
        if total_duration_min > 0 and total_m2_reported > 0:
            productivity_m2_h = round((total_m2_reported / (total_duration_min / 60)), 2)
        
        # Tempo médio por m²
        avg_time_per_m2 = 0
        if total_m2_reported > 0:
            avg_time_per_m2 = round(total_duration_min / total_m2_reported, 2)
        
        installer_data = {
            "installer_id": installer_id,
            "full_name": installer.get("full_name"),
            "branch": installer.get("branch"),
            "metrics": {
                "total_checkins": total_checkins,
                "completed_checkins": len(completed_checkins),
                "jobs_worked": jobs_worked,
                "total_duration_hours": round(total_duration_min / 60, 2),
                "total_m2_reported": total_m2_reported,
                "total_job_area_m2": round(total_job_area_m2, 2),
                "productivity_m2_h": productivity_m2_h,
                "avg_time_per_m2_min": avg_time_per_m2
            },
            "jobs": sorted(jobs_details, key=lambda x: x.get("job_area_m2", 0), reverse=True)[:20]
        }
        
        installer_report.append(installer_data)
    
    # Ordenar por produtividade
    installer_report.sort(key=lambda x: x["metrics"]["productivity_m2_h"], reverse=True)
    
    # Totais gerais
    total_area_all = sum(i["metrics"]["total_m2_reported"] for i in installer_report)
    total_hours_all = sum(i["metrics"]["total_duration_hours"] for i in installer_report)
    
    return {
        "summary": {
            "total_installers": len(installer_report),
            "total_area_m2_all": round(total_area_all, 2),
            "total_hours_all": round(total_hours_all, 2),
            "avg_productivity_m2_h": round(total_area_all / total_hours_all, 2) if total_hours_all > 0 else 0
        },
        "by_installer": installer_report
    }

@api_router.get("/metrics")
async def get_metrics(current_user: User = Depends(get_current_user)):
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    # Total jobs
    total_jobs = await db.jobs.count_documents({})
    completed_jobs = await db.jobs.count_documents({"status": "completed"})
    in_progress_jobs = await db.jobs.count_documents({"status": "in_progress"})
    pending_jobs = await db.jobs.count_documents({"status": "pending"})
    
    # Total checkins
    total_checkins = await db.checkins.count_documents({})
    completed_checkins = await db.checkins.count_documents({"status": "completed"})
    
    # Average duration
    completed_checkins_docs = await db.checkins.find({"status": "completed"}, {"duration_minutes": 1, "_id": 0}).to_list(1000)
    avg_duration = sum(c.get('duration_minutes', 0) for c in completed_checkins_docs) / len(completed_checkins_docs) if completed_checkins_docs else 0
    
    # Installers
    total_installers = await db.installers.count_documents({})
    
    return {
        "total_jobs": total_jobs,
        "completed_jobs": completed_jobs,
        "in_progress_jobs": in_progress_jobs,
        "pending_jobs": pending_jobs,
        "total_checkins": total_checkins,
        "completed_checkins": completed_checkins,
        "avg_duration_minutes": round(avg_duration, 2),
        "total_installers": total_installers
    }


@api_router.get("/reports/export")
async def export_reports(current_user: User = Depends(get_current_user)):
    """Export consolidated report to Excel"""
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    # Get all checkins with related data
    checkins = await db.checkins.find({}, {"_id": 0}).to_list(1000)
    jobs = await db.jobs.find({}, {"_id": 0}).to_list(1000)
    installers = await db.installers.find({}, {"_id": 0}).to_list(1000)
    
    # Create mapping dicts for faster lookup
    jobs_map = {job['id']: job for job in jobs}
    installers_map = {installer['id']: installer for installer in installers}
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Trabalhos"
    
    # Define styles
    header_fill = PatternFill(start_color="FF1F5A", end_color="FF1F5A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "ID do Job",
        "Nome do Job",
        "Cliente",
        "Área Total (m²)",
        "M² Instalado",
        "Instalador",
        "GPS Check-in (Lat)",
        "GPS Check-in (Long)",
        "GPS Check-out (Lat)",
        "GPS Check-out (Long)",
        "Data Check-in",
        "Data Check-out",
        "Tempo (min)",
        "Status",
        "Filial"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    row_num = 2
    for checkin in checkins:
        job = jobs_map.get(checkin.get('job_id'))
        installer = installers_map.get(checkin.get('installer_id'))
        
        if not job:
            continue
            
        ws.cell(row=row_num, column=1, value=job.get('id', '')).border = border
        ws.cell(row=row_num, column=2, value=job.get('title', '')).border = border
        ws.cell(row=row_num, column=3, value=job.get('client_name', '')).border = border
        ws.cell(row=row_num, column=4, value=job.get('area_m2', '')).border = border
        ws.cell(row=row_num, column=5, value=checkin.get('installed_m2', '')).border = border
        ws.cell(row=row_num, column=6, value=installer.get('full_name', '') if installer else '').border = border
        ws.cell(row=row_num, column=7, value=checkin.get('gps_lat', '')).border = border
        ws.cell(row=row_num, column=8, value=checkin.get('gps_long', '')).border = border
        ws.cell(row=row_num, column=9, value=checkin.get('checkout_gps_lat', '')).border = border
        ws.cell(row=row_num, column=10, value=checkin.get('checkout_gps_long', '')).border = border
        
        checkin_at = checkin.get('checkin_at')
        if isinstance(checkin_at, str):
            checkin_at = datetime.fromisoformat(checkin_at)
        ws.cell(row=row_num, column=11, value=checkin_at.strftime('%d/%m/%Y %H:%M') if checkin_at else '').border = border
        
        checkout_at = checkin.get('checkout_at')
        if isinstance(checkout_at, str):
            checkout_at = datetime.fromisoformat(checkout_at)
        ws.cell(row=row_num, column=12, value=checkout_at.strftime('%d/%m/%Y %H:%M') if checkout_at else '').border = border
        
        ws.cell(row=row_num, column=13, value=checkin.get('duration_minutes', '')).border = border
        ws.cell(row=row_num, column=14, value=checkin.get('status', '')).border = border
        ws.cell(row=row_num, column=15, value=job.get('branch', '')).border = border
        
        row_num += 1
    
    # Adjust column widths
    column_widths = {
        'A': 35, 'B': 30, 'C': 25, 'D': 15, 'E': 15,
        'F': 20, 'G': 18, 'H': 18, 'I': 18, 'J': 18,
        'K': 18, 'L': 18, 'M': 12, 'N': 15, 'O': 12
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename with current date
    filename = f"relatorio_trabalhos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/")
async def root():
    return {"message": "INDÚSTRIA VISUAL API", "status": "online"}

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()