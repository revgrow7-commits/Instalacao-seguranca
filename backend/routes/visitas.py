"""
Visitas Técnicas routes.
Gerencia o ciclo completo de visitas técnicas: criação, agendamento, cancelamento e consulta.
"""
from fastapi import APIRouter, HTTPException, Depends, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime, timezone
from io import BytesIO
import logging
import uuid
import base64

from db_supabase import db, upload_photo_to_storage
from security import get_current_user, require_role
from services.email import (
    send_vendedor_report_email,
    send_installer_invite_email,
    send_visita_notificacao_geral,
    send_relatorio_visita_concluida,
)
from models.user import User, UserRole
from models.visita import (
    VisitaCreate,
    VisitaUpdate,
    AgendarVisitaRequest,
    VisitaRelatorio,
    VisitaConfirmar,
    VisitaRejeitar,
    VisitaOut,
    VisitaStatus,
)
from routes.notifications import notify_visita_scheduled

router = APIRouter()
logger = logging.getLogger(__name__)


# ============ HELPER ============

def _parse_datetimes(doc: dict) -> dict:
    """Converte campos datetime de string ISO para datetime, para compatibilidade com Pydantic."""
    dt_fields = [
        "scheduled_date", "scheduled_time_end", "created_at", "updated_at",
        "relatorio_chegada", "relatorio_saida", "relatorio_enviado_em",
        "confirmado_em", "rejeitado_em",
    ]
    for field in dt_fields:
        val = doc.get(field)
        if val and isinstance(val, str):
            try:
                doc[field] = datetime.fromisoformat(val.replace("Z", "+00:00"))
            except ValueError:
                pass
    return doc


# Cache simples em memória: installer_id → full_name
_installer_name_cache: dict = {}

def _enrich_installer_name(doc: dict) -> dict:
    """Adiciona installer_name ao doc buscando na tabela installers pelo installer_id."""
    installer_id = doc.get("installer_id")
    if not installer_id:
        return doc
    if installer_id not in _installer_name_cache:
        rec = db.installers.find_one({"id": installer_id})
        _installer_name_cache[installer_id] = rec.get("full_name", "") if rec else ""
    doc["installer_name"] = _installer_name_cache[installer_id]
    return doc


# ============ ROUTES ============

@router.post("/visitas", response_model=VisitaOut)
async def create_visita(
    data: VisitaCreate,
    current_user: User = Depends(get_current_user),
):
    """Cria uma nova visita técnica. Somente admin/manager."""
    require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])

    now = datetime.now(timezone.utc).isoformat()
    visita_id = str(uuid.uuid4())

    # ── Resolver installer_id por email quando instalador externo sem id local ──
    effective_installer_id = data.installer_id
    if data.installer_email and not effective_installer_id:
        local_installer = db.installers.find_one({"email": data.installer_email.lower()})
        if not local_installer:
            # Tentar via tabela users (instaladores cujo user_id aponta para users.email)
            user_rec = db.users.find_one({"email": data.installer_email.lower()})
            if user_rec:
                local_installer = db.installers.find_one({"user_id": user_rec["id"]})
        if local_installer:
            effective_installer_id = local_installer["id"]

    doc = {
        "id": visita_id,
        "titulo": data.titulo or "VISITA TÉCNICA",
        "client_name": data.client_name,
        "client_address": data.client_address,
        "branch": data.branch,
        "installer_id": effective_installer_id,
        "installer_nome": data.installer_nome,
        "installer_email": data.installer_email,
        "scheduled_date": data.scheduled_date.isoformat() if data.scheduled_date else None,
        "scheduled_time_end": data.scheduled_time_end.isoformat() if data.scheduled_time_end else None,
        "valor_por_km": data.valor_por_km if data.valor_por_km is not None else 1.50,
        "km_ida": data.km_ida,
        "km_volta": data.km_volta,
        "status": VisitaStatus.AGUARDANDO.value,
        "observacoes_admin": data.observacoes_admin,
        "relatorio_descricao": None,
        "relatorio_situacao": None,
        "relatorio_fotos": [],
        "relatorio_assinatura_confirmada": False,
        "relatorio_chegada": None,
        "relatorio_saida": None,
        "relatorio_enviado_em": None,
        "created_by": current_user.id,
        "created_at": now,
        "updated_at": now,
        # Novos campos — expansão VT
        "job_id": data.job_id,
        "vendedor_nome": data.vendedor_nome,
        "vendedor_email": data.vendedor_email,
        "tipos_servico": data.tipos_servico or [],
        "ferramentas": data.ferramentas or [],
        "remocao_prevista_os": data.remocao_prevista_os or False,
        "remocao_a_realizar": data.remocao_a_realizar or False,
        "altura_estimada_m": data.altura_estimada_m,
        "nivel_dificuldade": data.nivel_dificuldade,
        "aprovacao_status": data.aprovacao_status or "PENDENTE",
    }

    db.visitas_tecnicas.insert_one(doc)

    saved = db.visitas_tecnicas.find_one({"id": visita_id})
    if not saved:
        raise HTTPException(status_code=500, detail="Erro ao salvar visita técnica")

    # ── Notifica todos os managers (vendedores) e instaladores ───────────────
    try:
        send_visita_notificacao_geral(saved)
    except Exception as exc:
        logger.error("Erro ao enviar notificações de nova VT: %s", exc)

    return VisitaOut(**_parse_datetimes(_enrich_installer_name(saved)))


@router.get("/visitas", response_model=List[VisitaOut])
async def list_visitas(
    status: Optional[str] = Query(None),
    installer_id: Optional[str] = Query(None),
    branch: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    # Aliases para compatibilidade com clientes que enviam start_date/end_date
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    vendedor_email: Optional[str] = Query(None),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(get_current_user),
):
    """
    Lista visitas técnicas com filtros opcionais e paginação.
    Admin/manager vê todas; installer só vê as suas.
    Aceita date_from/date_to ou start_date/end_date (aliases).
    """
    query: dict = {}

    if current_user.role == UserRole.INSTALLER:
        # visitas_tecnicas.installer_id aponta para installers.id, não users.id
        # precisa fazer o mapeamento users.id → installers.id via installers.user_id
        installer_rec = db.installers.find_one({"user_id": current_user.id})
        if not installer_rec:
            return []  # instalador sem registro na tabela installers
        query["installer_id"] = installer_rec["id"]
    else:
        if installer_id:
            query["installer_id"] = installer_id

    if status:
        query["status"] = status

    if branch:
        query["branch"] = branch

    _date_from = date_from or start_date
    _date_to = date_to or end_date
    if _date_from or _date_to:
        date_filter: dict = {}
        if _date_from:
            date_filter["$gte"] = _date_from
        if _date_to:
            date_filter["$lte"] = _date_to
        query["scheduled_date"] = date_filter

    if vendedor_email:
        query["vendedor_email"] = {"$regex": vendedor_email}

    docs = db.visitas_tecnicas.find(
        query,
        {"_id": 0},
        sort=[("created_at", -1)],
        limit=limit,
        skip=offset,
    )

    return [VisitaOut(**_parse_datetimes(_enrich_installer_name(d))) for d in (docs or [])]


@router.get("/visitas/reports/excel")
async def export_visitas_excel(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    # Aliases para compatibilidade
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    installer_id: Optional[str] = Query(None),
    branch: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
):
    """
    Gera planilha Excel com visitas técnicas filtradas (28 colunas).
    Aceita date_from/date_to ou start_date/end_date (aliases).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])

    query: dict = {}
    if installer_id:
        query["installer_id"] = installer_id
    if branch:
        query["branch"] = branch

    _date_from = date_from or start_date
    _date_to = date_to or end_date
    if _date_from or _date_to:
        date_filter: dict = {}
        if _date_from:
            date_filter["$gte"] = _date_from
        if _date_to:
            date_filter["$lte"] = _date_to
        query["scheduled_date"] = date_filter

    docs_raw = db.visitas_tecnicas.find(query, {"_id": 0}, sort=[("created_at", -1)]) or []
    docs = [_enrich_installer_name(dict(d)) for d in docs_raw]

    wb = Workbook()
    ws = wb.active
    ws.title = "Visitas Técnicas"

    header_fill = PatternFill(start_color="FF1F5A", end_color="FF1F5A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Nº VT", "Cliente", "Endereço", "Filial", "Instalador",
        "Data Agendada", "KM Ida", "KM Volta", "Valor/KM", "Total (R$)",
        "Vendedor", "Tipo de Serviço", "Remoção Prevista", "Remoção Realizada",
        "Altura (m)", "Ferramentas", "Nível Dificuldade", "Status Aprovação", "Status",
        # novas
        "Rel. Situação", "Rel. Enviado Em", "Estacionamento", "Ponto Energia",
        "Tipo Superfície", "Forma Instalação", "EPI Altura", "Escada", "Andaime Torres",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    _nivel_labels = {
        1: "Nível 1 - Simples",
        2: "Nível 2 - Moderado",
        3: "Nível 3 - Complexo",
        4: "Nível 4 - Crítico",
    }
    _aprovacao_labels = {
        "PENDENTE": "Pendente",
        "APROVADO": "Aprovado",
        "NAO_APROVADO": "Não aprovado",
    }

    def _fmt_dt(val):
        if not val:
            return ""
        if isinstance(val, str):
            try:
                val = datetime.fromisoformat(val.replace("Z", "+00:00"))
            except (ValueError, TypeError):
                return val
        return val.strftime("%d/%m/%Y %H:%M") if isinstance(val, datetime) else str(val)

    for row_num, doc in enumerate(docs, start=2):
        scheduled_raw = doc.get("scheduled_date")
        if isinstance(scheduled_raw, str):
            try:
                scheduled_raw = datetime.fromisoformat(scheduled_raw.replace("Z", "+00:00"))
                scheduled_raw = scheduled_raw.strftime("%d/%m/%Y %H:%M")
            except (ValueError, TypeError):
                pass

        tipos = doc.get("tipos_servico") or []
        ferramentas = doc.get("ferramentas") or []
        nivel = doc.get("nivel_dificuldade")
        aprovacao = doc.get("aprovacao_status") or "PENDENTE"

        row = [
            doc.get("numero_vt") or "",
            doc.get("client_name", ""),
            doc.get("client_address", ""),
            doc.get("branch", ""),
            doc.get("installer_name") or doc.get("installer_id") or "",
            scheduled_raw or "",
            doc.get("km_ida") or "",
            doc.get("km_volta") or "",
            doc.get("valor_por_km", 1.50),
            doc.get("valor_total") or "",
            doc.get("vendedor_nome") or "",
            ", ".join(tipos) if isinstance(tipos, list) else tipos or "",
            "Sim" if doc.get("remocao_prevista_os") else "Não",
            "Sim" if doc.get("remocao_a_realizar") else "Não",
            doc.get("altura_estimada_m") or "",
            ", ".join(ferramentas) if isinstance(ferramentas, list) else ferramentas or "",
            _nivel_labels.get(nivel, "") if nivel is not None else "",
            _aprovacao_labels.get(aprovacao, aprovacao),
            doc.get("status", ""),
            # novas
            doc.get("relatorio_situacao") or "",
            _fmt_dt(doc.get("relatorio_enviado_em")),
            "Sim" if doc.get("tem_estacionamento") else ("Não" if doc.get("tem_estacionamento") is False else ""),
            "Sim" if doc.get("tem_ponto_energia") else ("Não" if doc.get("tem_ponto_energia") is False else ""),
            ", ".join(doc.get("tipo_superficie") or []),
            ", ".join(doc.get("forma_instalacao") or []),
            "Sim" if doc.get("epi_altura") else ("Não" if doc.get("epi_altura") is False else ""),
            doc.get("escada_tamanho") or "",
            doc.get("andaime_torres") or "",
        ]

        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=value).border = border

    column_widths = {
        "A": 14, "B": 30, "C": 35, "D": 12, "E": 30,
        "F": 18, "G": 12, "H": 12, "I": 12, "J": 14,
        "K": 16, "L": 28, "M": 18, "N": 18,
        "O": 12, "P": 30, "Q": 22, "R": 18, "S": 14,
        # novas
        "T": 18, "U": 20, "V": 14, "W": 14, "X": 28, "Y": 28, "Z": 12, "AA": 18, "AB": 14,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"visitas_tecnicas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/visitas/{visita_id}", response_model=VisitaOut)
async def get_visita(
    visita_id: str,
    current_user: User = Depends(get_current_user),
):
    """Busca uma visita técnica por ID. Installer só acessa a sua."""
    doc = db.visitas_tecnicas.find_one({"id": visita_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Visita técnica não encontrada")

    if current_user.role == UserRole.INSTALLER:
        # visitas_tecnicas.installer_id aponta para installers.id, não users.id
        # precisa fazer o mapeamento users.id → installers.id via installers.user_id
        installer_rec = db.installers.find_one({"user_id": current_user.id})
        if not installer_rec or doc.get("installer_id") != installer_rec["id"]:
            raise HTTPException(status_code=403, detail="Você não tem acesso a esta visita técnica")

    return VisitaOut(**_parse_datetimes(_enrich_installer_name(doc)))


@router.patch("/visitas/{visita_id}", response_model=VisitaOut)
async def update_visita(
    visita_id: str,
    data: VisitaUpdate,
    current_user: User = Depends(get_current_user),
):
    """Atualiza campos de uma visita técnica. Somente admin/manager. Bloqueado se CONCLUIDA."""
    require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])

    doc = db.visitas_tecnicas.find_one({"id": visita_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Visita técnica não encontrada")

    if doc.get("status") == VisitaStatus.CONCLUIDA.value:
        raise HTTPException(status_code=400, detail="Não é possível editar uma visita técnica já concluída")

    update_fields = data.model_dump(exclude_unset=True)
    if not update_fields:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")

    # Serializa datetimes para ISO string
    for field in ("scheduled_date", "scheduled_time_end", "relatorio_chegada", "relatorio_saida", "relatorio_enviado_em"):
        if field in update_fields and isinstance(update_fields[field], datetime):
            update_fields[field] = update_fields[field].isoformat()

    # Converte enum para valor string
    if "status" in update_fields and isinstance(update_fields["status"], VisitaStatus):
        update_fields["status"] = update_fields["status"].value

    update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()

    db.visitas_tecnicas.update_one({"id": visita_id}, {"$set": update_fields})

    updated = db.visitas_tecnicas.find_one({"id": visita_id}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=500, detail="Erro ao recuperar visita técnica após atualização")

    return VisitaOut(**_parse_datetimes(_enrich_installer_name(updated)))


@router.post("/visitas/{visita_id}/agendar", response_model=VisitaOut)
async def agendar_visita(
    visita_id: str,
    data: AgendarVisitaRequest,
    current_user: User = Depends(get_current_user),
):
    """
    Atribui instalador e define data/hora da visita técnica.
    Somente admin/manager. Retorna 400 se status=CANCELADA.
    """
    require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])

    doc = db.visitas_tecnicas.find_one({"id": visita_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Visita técnica não encontrada")

    if doc.get("status") == VisitaStatus.CANCELADA.value:
        raise HTTPException(status_code=400, detail="Não é possível agendar uma visita técnica cancelada")

    update_fields = {
        "installer_id": data.installer_id,
        "scheduled_date": data.scheduled_date.isoformat(),
        "scheduled_time_end": data.scheduled_time_end.isoformat() if data.scheduled_time_end else None,
        "status": VisitaStatus.AGUARDANDO_CONFIRMACAO.value,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    if data.observacoes_admin is not None:
        update_fields["observacoes_admin"] = data.observacoes_admin

    db.visitas_tecnicas.update_one({"id": visita_id}, {"$set": update_fields})

    updated = db.visitas_tecnicas.find_one({"id": visita_id}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=500, detail="Erro ao recuperar visita técnica após agendamento")

    # Invalida cache do nome para este instalador (pode ter mudado)
    _installer_name_cache.pop(data.installer_id, None)

    # Push notification para o instalador atribuído (silencioso em caso de falha)
    try:
        await notify_visita_scheduled(
            visita_id=visita_id,
            installer_id=data.installer_id,
            visita_title=updated.get("titulo") or "Visita Técnica",
            client_name=updated.get("client_name") or "",
        )
    except Exception as e:
        logger.warning(f"Falha ao notificar instalador {data.installer_id} sobre VT {visita_id}: {e}")

    return VisitaOut(**_parse_datetimes(_enrich_installer_name(updated)))


# Campos que podem ser editados pelo instalador na confirmação e que entram no snapshot do plano admin.
_CONFIRMAR_FIELDS = (
    "km_ida",
    "km_volta",
    "altura_estimada_m",
    "nivel_dificuldade",
    "ferramentas",
    "remocao_a_realizar",
    "tipos_servico",
    # checklist de vistoria
    "tem_estacionamento",
    "restricao_horario_inicio",
    "restricao_horario_fim",
    "tipo_superficie",
    "tipo_superficie_outro",
    "condicao_superficie",
    "material_remocao",
    "tem_ponto_energia",
    "medida_largura_m",
    "medida_altura_m",
    "forma_instalacao",
    "epi_altura",
    "escada_tamanho",
    "andaime_torres",
)


def _check_installer_owns_visita(doc: dict, current_user: User) -> None:
    """Garante que o usuário atual é o instalador atribuído à visita. Levanta 403 caso contrário."""
    if current_user.role != UserRole.INSTALLER:
        raise HTTPException(status_code=403, detail="Somente instaladores podem executar esta ação")
    installer_rec = db.installers.find_one({"user_id": current_user.id})
    if not installer_rec or doc.get("installer_id") != installer_rec["id"]:
        raise HTTPException(status_code=403, detail="Você não está atribuído a esta visita técnica")


@router.post("/visitas/{visita_id}/confirmar", response_model=VisitaOut)
async def confirmar_visita(
    visita_id: str,
    data: VisitaConfirmar,
    current_user: User = Depends(get_current_user),
):
    """
    Instalador confirma a visita técnica, opcionalmente ajustando campos planejados.
    Snapshot do plano admin é gravado em planejado_snapshot e status passa para EM_VISITA.
    """
    doc = db.visitas_tecnicas.find_one({"id": visita_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Visita técnica não encontrada")

    _check_installer_owns_visita(doc, current_user)

    if doc.get("status") != VisitaStatus.AGUARDANDO_CONFIRMACAO.value:
        raise HTTPException(
            status_code=400,
            detail="Visita técnica não está aguardando confirmação",
        )

    # Snapshot dos valores planejados (somente para os campos editáveis pelo instalador)
    snapshot = {field: doc.get(field) for field in _CONFIRMAR_FIELDS}

    edits = data.model_dump(exclude_unset=True)
    update_fields: dict = {
        "planejado_snapshot": snapshot,
        "confirmado_em": datetime.now(timezone.utc).isoformat(),
        "confirmado_por": current_user.id,
        "status": VisitaStatus.EM_VISITA.value,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    update_fields.update(edits)

    db.visitas_tecnicas.update_one({"id": visita_id}, {"$set": update_fields})

    updated = db.visitas_tecnicas.find_one({"id": visita_id}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=500, detail="Erro ao recuperar visita técnica após confirmação")

    return VisitaOut(**_parse_datetimes(_enrich_installer_name(updated)))


@router.post("/visitas/{visita_id}/rejeitar", response_model=VisitaOut)
async def rejeitar_visita(
    visita_id: str,
    data: VisitaRejeitar,
    current_user: User = Depends(get_current_user),
):
    """
    Instalador rejeita a visita técnica com motivo (>= 10 chars).
    Status volta para AGUARDANDO e installer_id é limpo para reagendamento pelo admin.
    """
    doc = db.visitas_tecnicas.find_one({"id": visita_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Visita técnica não encontrada")

    _check_installer_owns_visita(doc, current_user)

    if doc.get("status") != VisitaStatus.AGUARDANDO_CONFIRMACAO.value:
        raise HTTPException(
            status_code=400,
            detail="Visita técnica não está aguardando confirmação",
        )

    update_fields = {
        "rejeitado_em": datetime.now(timezone.utc).isoformat(),
        "rejeitado_motivo": data.motivo,
        "status": VisitaStatus.AGUARDANDO.value,
        "installer_id": None,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    db.visitas_tecnicas.update_one({"id": visita_id}, {"$set": update_fields})

    updated = db.visitas_tecnicas.find_one({"id": visita_id}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=500, detail="Erro ao recuperar visita técnica após rejeição")

    return VisitaOut(**_parse_datetimes(_enrich_installer_name(updated)))


@router.post("/visitas/{visita_id}/cancelar", response_model=VisitaOut)
async def cancelar_visita(
    visita_id: str,
    current_user: User = Depends(get_current_user),
):
    """Cancela uma visita técnica. Somente admin/manager. Bloqueado se já CONCLUIDA."""
    require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])

    doc = db.visitas_tecnicas.find_one({"id": visita_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Visita técnica não encontrada")

    if doc.get("status") == VisitaStatus.CONCLUIDA.value:
        raise HTTPException(status_code=400, detail="Não é possível cancelar uma visita técnica já concluída")

    db.visitas_tecnicas.update_one(
        {"id": visita_id},
        {"$set": {
            "status": VisitaStatus.CANCELADA.value,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
    )

    updated = db.visitas_tecnicas.find_one({"id": visita_id}, {"_id": 0})
    if not updated:
        raise HTTPException(status_code=500, detail="Erro ao recuperar visita técnica após cancelamento")

    return VisitaOut(**_parse_datetimes(_enrich_installer_name(updated)))


@router.post("/visitas/{visita_id}/enviar-email")
async def enviar_email_relatorio(
    visita_id: str,
    current_user: User = Depends(get_current_user),
):
    """Envia o relatório completo da visita por email ao vendedor e/ou instalador. Somente admin/manager."""
    require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])

    doc = db.visitas_tecnicas.find_one({"id": visita_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Visita técnica não encontrada")

    sent_to = send_relatorio_visita_concluida(doc)

    if not sent_to:
        raise HTTPException(
            status_code=400,
            detail="Nenhum email de destinatário configurado na visita (vendedor_email / installer_email).",
        )

    return {"success": True, "sent_to": sent_to}


@router.post("/visitas/{visita_id}/relatorio", response_model=VisitaOut)
async def enviar_relatorio(
    visita_id: str,
    km_ida: Optional[float] = Form(None),
    km_volta: float = Form(0.0, ge=0),
    km_rodados: Optional[float] = Form(None),
    descricao: str = Form(...),
    situacao: str = Form(...),
    assinatura_confirmada: bool = Form(False),
    chegada: Optional[str] = Form(None),
    saida: Optional[str] = Form(None),
    fotos: List[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
):
    """
    Envia o relatório de uma visita técnica (multipart/form-data).
    Somente o instalador atribuído pode enviar. Requer mínimo 1 foto.
    Status muda para CONCLUIDA.
    """
    # 1. Buscar visita
    doc = db.visitas_tecnicas.find_one({"id": visita_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Visita técnica não encontrada")

    # 2. Verificar permissão: somente o instalador atribuído
    if current_user.role != UserRole.INSTALLER:
        raise HTTPException(status_code=403, detail="Somente instaladores podem enviar relatórios")

    # visitas_tecnicas.installer_id aponta para installers.id, não users.id
    # precisa fazer o mapeamento users.id → installers.id via installers.user_id
    installer_rec = db.installers.find_one({"user_id": current_user.id})
    if not installer_rec or doc.get("installer_id") != installer_rec["id"]:
        raise HTTPException(status_code=403, detail="Você não está atribuído a esta visita técnica")

    # 3. Verificar status — só aceita relatório quando a visita está EM_VISITA
    current_status = doc.get("status")
    if current_status == VisitaStatus.CONCLUIDA.value:
        raise HTTPException(status_code=400, detail="Relatório já enviado")
    if current_status == VisitaStatus.CANCELADA.value:
        raise HTTPException(status_code=400, detail="Visita cancelada")
    if current_status != VisitaStatus.EM_VISITA.value:
        raise HTTPException(
            status_code=400,
            detail="Confirme a visita antes de enviar o relatório (status precisa ser EM_VISITA)",
        )

    # 4. Resolver km — campo principal: km_ida; retrocompat: km_rodados (antigos clientes)
    effective_km_ida = km_ida if km_ida is not None else km_rodados
    if not effective_km_ida or effective_km_ida <= 0:
        raise HTTPException(status_code=422, detail="km_ida deve ser maior que zero")

    # 5. Validar fotos
    if not fotos or len(fotos) < 1:
        raise HTTPException(status_code=400, detail="Mínimo 1 foto obrigatória")

    # 6. Upload de fotos — padrão idêntico ao routes/checkins.py
    lista_de_urls: list = []
    for i, foto in enumerate(fotos):
        foto_bytes = await foto.read()
        base64_str = base64.b64encode(foto_bytes).decode("utf-8")
        url = upload_photo_to_storage(
            base64_str,
            f"visitas-tecnicas/{visita_id}/{i}.jpg",
            bucket="checkin-photos",
        )
        lista_de_urls.append({"index": i, "url": url or base64_str})

    # 7. Converter chegada/saida ISO → string para armazenamento
    chegada_iso: Optional[str] = None
    if chegada:
        try:
            chegada_iso = datetime.fromisoformat(chegada.replace("Z", "+00:00")).isoformat()
        except ValueError:
            chegada_iso = chegada

    saida_iso: Optional[str] = None
    if saida:
        try:
            saida_iso = datetime.fromisoformat(saida.replace("Z", "+00:00")).isoformat()
        except ValueError:
            saida_iso = saida

    # 8. Montar campos de atualização (valor_total é GENERATED ALWAYS — não gravar)
    update_fields = {
        "km_ida": effective_km_ida,
        "km_volta": km_volta,
        "status": VisitaStatus.CONCLUIDA.value,
        "relatorio_descricao": descricao,
        "relatorio_situacao": situacao,
        "relatorio_fotos": lista_de_urls,
        "relatorio_assinatura_confirmada": assinatura_confirmada,
        "relatorio_chegada": chegada_iso,
        "relatorio_saida": saida_iso,
        "relatorio_enviado_em": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    # 9. Persistir
    db.visitas_tecnicas.update_one({"id": visita_id}, {"$set": update_fields})

    # 10. Retornar visita atualizada
    saved = db.visitas_tecnicas.find_one({"id": visita_id}, {"_id": 0})
    if not saved:
        raise HTTPException(status_code=500, detail="Erro ao recuperar visita técnica após atualização")

    return VisitaOut(**_parse_datetimes(_enrich_installer_name(saved)))
